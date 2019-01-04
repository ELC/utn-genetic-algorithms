"""Reports"""

import pandas as pd
import time
from openpyxl import load_workbook
from exercise1.util.graphic import graphics

pd.set_option('display.max_rows', 1500)
pd.set_option('display.max_columns', 700)
pd.set_option('display.width', 1500)
pd.set_option('precision', 12)


class Report():
    FileManager = None
    Controller = None
    Population = None


    @classmethod
    def generations_report(cls, execution_time):
        """Show the generations report"""
        data = cls.FileManager.load_results()
        datas = cls._get_array_data(data)
        labels = ("1-Máx",
                  "2-Mín",
                  "3-Promedio",
                  "5-Rango",
                  "4-Total",
                  "6-Cromosoma",
                  "7-Estable")
        data_pd = {i: pd.Series(j) for i, j in zip(labels, datas)}
        data_frame = pd.DataFrame(data_pd)
        data_frame.index.names = ['G']
        # data_frame = data_frame.reindex(data_frame.index.rename("Generacion"))
        cls.write_csv(data_frame, execution_time)
        cls.write_excel(data_frame, execution_time)
        return data_frame

    @classmethod
    def write_csv(cls, df, execution_time):
        name = str(cls.FileManager.get_settings_id())[:5]
        filename = name + '.csv'
        if cls.is_empty(filename):
            settings = cls.FileManager.load_settings()
            pd.DataFrame(list(settings.items())).to_csv(
                filename, mode="a+", index=False)
            cls.add_empty_line(filename)
        df.to_csv(filename, mode="a+")
        cls.add_empty_line(filename)
        cls.add_elapsed_time(filename, execution_time)
        cls.add_empty_line(filename)
        cls.add_empty_line(filename)

    @classmethod
    def add_empty_line(cls, filename):
        with open(filename, "a+") as handler:
            handler.write('\n')

    @classmethod
    def add_elapsed_time(cls, filename, execution_time):
        with open(filename, "a+") as handler:
            handler.write(str(execution_time))

    @classmethod
    def is_empty(cls, filename):
        try:
            with open(filename) as _:
                pass
            return False
        except FileNotFoundError:
            return True

    @classmethod
    def write_excel(cls, df, execution_time):
        sheet_name = str(cls.FileManager.get_settings_id())[1:5]
        filename = 'resultados.xlsx'
        writer = pd.ExcelWriter(filename, engine='openpyxl')

        try:
            writer.book = load_workbook(filename)
            writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
        except:
            pass

        settings = cls.FileManager.load_settings()

        try:
            startrow = writer.book[sheet_name].max_row
        except:
            startrow = -2

        settings_df = pd.DataFrame(list(settings.items()))
        settings_df.to_excel(writer, sheet_name=sheet_name, index=False, startcol=8)

        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=startrow + 2)

        df.tail(1).to_excel(writer, sheet_name=sheet_name, index=False, header=None, startcol=11, startrow=startrow + 2)
        
        time_df = pd.DataFrame([("Time", execution_time)])
        time_df.to_excel(writer, sheet_name=sheet_name, index=False, header=None, startcol=19, startrow=startrow + 2)

        writer.save()
        writer.close()

    @classmethod
    def solution_report(cls):
        """Show the final solution infered from the generations"""
        data = cls.FileManager.load_results()
        last_population = data[-1]
        solution = cls.get_maximum(last_population)
        return solution

    @classmethod
    def _get_array_data(cls, data):
        maximums = [cls.get_maximum(population) for population in data]
        minimums = [cls.get_minimum(population) for population in data]
        averages = [cls.get_average(population) for population in data]
        ranges = [cls.get_range(population) for population in data]
        total = [cls.get_sum(population) for population in data]
        chromosome = [cls.Population.get_fittest_chromosome(population) for population in data]
        stationary = cls._get_stationary(data)
        return (maximums, minimums, averages, ranges, total, chromosome, stationary)

    @classmethod
    def graphic_min_max_mean(cls):
        data = cls.FileManager.load_results()
        maximums, minimums, averages, *_ = cls._get_array_data(data)
        graphics(datas=(maximums, averages, minimums),
                labels=["Máximo", "Promedio", "Minimo"])

    @classmethod
    def graphic_range(cls):
        data = cls.FileManager.load_results()
        _, _, _, ranges, *_ = cls._get_array_data(data)
        graphics(datas=[ranges], labels=["Rangos"])

    @classmethod
    def graphic_t(cls):
        data = cls.FileManager.load_results()
        _, _, _, _, total, *_ = cls._get_array_data(data)
        graphics(datas=[total], labels=["Totales"])

    @classmethod
    def get_maximum(cls, population):
        return max(cls.Population.get_targets(population))

    @classmethod
    def get_minimum(cls, population):
        return min(cls.Population.get_targets(population))

    @classmethod
    def get_sum(cls, population):
        print(cls.Population.get_targets(population))
        return sum(cls.Population.get_targets(population))

    @classmethod
    def get_average(cls, population):
        return cls.get_sum(population) / len(population)

    @classmethod
    def get_range(cls, population):
        return cls.get_maximum(population) - cls.get_minimum(population)

    @classmethod
    def _get_stationary(cls, data):
        aux = ""
        stationary = []
        for generation, population in enumerate(data):
            max_chromosome = cls.Population.get_fittest_chromosome(population)
            if max_chromosome != aux:
                aux = max_chromosome
                last_value = generation + 1
            stationary.append(last_value)
        return stationary
