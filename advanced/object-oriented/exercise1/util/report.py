"""Reports"""

from exercise1.logic.population_manager import PopulationController
import pandas as pd
import time
from openpyxl import load_workbook
from exercise1.data.filemanager import FileManager as Settings
from exercise1.logic.controller import Controller
from exercise1.util.graphic import graphics

pd.set_option('display.max_rows', 1500)
pd.set_option('display.max_columns', 700)
pd.set_option('display.width', 1500)
pd.set_option('precision', 12)


class Report():

    @classmethod
    def full_report(cls):
        return cls.generations_report(), cls.solution_report()

    @classmethod
    def generations_report(cls):
        """Show the generations report"""
        data = PopulationController.load_populations()
        datas = cls._get_array_data(data)
        labels = ("1-Máx",
                  "2-Mín",
                  "3-Promedio",
                  "5-Rango",
                  "4-Total",
                  "6-Cromosoma",
                  "7-Estable")
        data_pd = {i: pd.Series(j) for i, j in zip(labels, datas)}
        data_frame = pd.DataFrame(data_pd)
        data_frame.index.names = ['G']
        # data_frame = data_frame.reindex(data_frame.index.rename("Generacion"))
        cls.write_csv(data_frame)
        cls.write_excel(data_frame)
        return data_frame

    @classmethod
    def write_csv(cls, df):
        name = str(Settings.get_settings_id())[:5]
        filename = name + '.csv'
        if cls.is_empty(filename):
            settings = Settings.load_settings()
            pd.DataFrame(list(settings.items())).to_csv(
                filename, mode="a+", index=False)
            cls.add_empty_line(filename)
        df.to_csv(filename, mode="a+")
        cls.add_empty_line(filename)
        cls.add_elapsed_time(filename)
        cls.add_empty_line(filename)
        cls.add_empty_line(filename)

    @classmethod
    def add_empty_line(cls, filename):
        with open(filename, "a+") as handler:
            handler.write('\n')

    @classmethod
    def add_elapsed_time(cls, filename):
        with open(filename, "a+") as handler:
            handler.write(str(Controller.get_execution_time()))

    @classmethod
    def is_empty(cls, filename):
        try:
            with open(filename) as _:
                pass
            return False
        except FileNotFoundError:
            return True

    @classmethod
    def write_excel(cls, df):
        sheet_name = str(Settings.get_settings_id())[1:5]
        filename = 'resultados.xlsx'
        writer = pd.ExcelWriter(filename, engine='openpyxl')

        try:
            writer.book = load_workbook(filename)
            writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
        except:
            pass

        settings = Settings.load_settings()

        try:
            startrow = writer.book[sheet_name].max_row
        except:
            startrow = -2

        settings_df = pd.DataFrame(list(settings.items()))
        settings_df.to_excel(writer, sheet_name=sheet_name, index=False, startcol=8)

        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=startrow + 2)

        df.tail(1).to_excel(writer, sheet_name=sheet_name, index=False, header=None, startcol=11, startrow=startrow + 2)
        
        time_df = pd.DataFrame([("Time", Controller.get_execution_time())])
        time_df.to_excel(writer, sheet_name=sheet_name, index=False, header=None, startcol=19, startrow=startrow + 2)

        writer.save()
        writer.close()

    @classmethod
    def solution_report(cls):
        """Show the final solution infered from the generations"""
        data = PopulationController.load_populations()
        last_population = data[-1]
        solution = last_population.get_maximum()
        return solution

    @classmethod
    def _get_array_data(cls, data):
        maximums = [population.get_maximum() for population in data]
        minimums = [population.get_minimum() for population in data]
        averages = [population.get_average() for population in data]
        ranges = [population.get_range() for population in data]
        total = [population.get_sum() for population in data]
        chromosome = [population.get_max_gene_string() for population in data]
        stationary = cls._get_stationary(data)
        return (maximums, minimums, averages, ranges, total, chromosome, stationary)

    @classmethod
    def graphic_min_max_mean(cls):
        data = PopulationController.load_populations()
        maximums, minimums, averages, *_ = cls._get_array_data(data)
        graphics(datas=(maximums, averages, minimums),
                labels=["Máximo", "Promedio", "Minimo"])

    @classmethod
    def graphic_range(cls):
        data = PopulationController.load_populations()
        _, _, _, ranges, *_ = cls._get_array_data(data)
        graphics(datas=[ranges], labels=["Rangos"])

    @classmethod
    def graphic_t(cls):
        data = PopulationController.load_populations()
        _, _, _, _, total, *_ = cls._get_array_data(data)
        graphics(datas=[total], labels=["Totales"])


    @classmethod
    def _get_stationary(cls, data):
        aux = ""
        stationary = []
        for population in data:
            if population.get_max_gene_string() != aux:
                aux = population.get_max_gene_string()
                last_value = int(population.get_generation())-1
            stationary.append(last_value)
        return stationary
